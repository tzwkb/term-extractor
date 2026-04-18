#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OpenAI LLM批处理术语抽取处理器
支持并发批量文本处理、任务监控和结果处理
"""

import json
import time
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

try:
    from openai import OpenAI
    import tiktoken
except ImportError:
    print("请安装OpenAI库: pip install openai tiktoken")
    raise


class LLMProcessor:
    """OpenAI LLM批处理处理器"""
    
    def __init__(self,
                 api_key: str,
                 base_url: str = "https://api.openai.com/v1",
                 base_dir: str = "batch_results"):
        """
        初始化批处理器

        Args:
            api_key: OpenAI API密钥
            base_url: API端点URL
            base_dir: 结果存储目录
        """
        self.client = OpenAI(
            api_key=api_key,
            base_url=base_url,
            timeout=60.0,  # 设置60秒超时
            max_retries=3   # 最大重试3次
        )
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(exist_ok=True)

        # 配置日志和并发控制
        self._setup_logging()
        self._setup_concurrency_control()
    
    # =============================================================================
    # 初始化和配置
    # =============================================================================
    
    def _setup_logging(self):
        """设置日志配置"""
        from config import LOGGING_CONFIG

        log_file = self.base_dir / "llm_processor.log"

        # 获取配置的日志级别
        log_level = getattr(logging, LOGGING_CONFIG.get("level", "INFO").upper())

        handlers = []
        if LOGGING_CONFIG.get("file_enabled", True):
            handlers.append(logging.FileHandler(log_file, encoding='utf-8'))
        if LOGGING_CONFIG.get("console_enabled", True):
            handlers.append(logging.StreamHandler())

        logging.basicConfig(
            level=log_level,
            format=LOGGING_CONFIG.get("format", '%(asctime)s - %(name)s - %(levelname)s - %(message)s'),
            handlers=handlers,
            force=True  # 强制重新配置日志系统
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"📋 日志级别: {LOGGING_CONFIG.get('level', 'INFO')}")
    
    def _setup_concurrency_control(self):
        """设置并发控制"""
        # 降低并发数以减少连接问题
        self.semaphore = threading.Semaphore(5)  # 最大5个并发请求
        self.logger.info("LLM处理器初始化完成")
    
    # =============================================================================
    # Token计算
    # =============================================================================
    
    def count_tokens(self, text: str, model: str = "gpt-4") -> int:
        """计算文本的token数量"""
        try:
            # 尝试获取模型特定的编码器
            encoding = tiktoken.encoding_for_model(model)
            return len(encoding.encode(text))
        except Exception:
            # 如果模型不支持，使用通用编码器
            encoding = tiktoken.get_encoding("cl100k_base")
            return len(encoding.encode(text))
    
    # =============================================================================
    # 单文本处理
    # =============================================================================
    
    def process_single_text(self, 
                          text: str,
                          custom_id: str,
                          system_prompt: str = None,
                          user_prompt_template: str = None,
                          model: str = "gpt-4-turbo-preview",
                          temperature: float = 0.1,
                          max_tokens: int = 4096,
                          source_file: str = None) -> Dict[str, Any]:
        """
        处理单个文本
        
        Args:
            text: 要处理的文本
            custom_id: 自定义ID
            system_prompt: 系统提示词
            user_prompt_template: 用户提示词模板
            model: 模型名称
            temperature: 温度参数
            max_tokens: 最大输出token数
            source_file: 来源文件名
            
        Returns:
            处理结果字典
        """
        with self.semaphore:  # 控制并发数
            try:
                # 验证必要参数
                if system_prompt is None:
                    raise ValueError("system_prompt is required")
                
                if user_prompt_template is None:
                    raise ValueError("user_prompt_template is required")
                
                # 构建API调用参数
                api_params = self._build_api_params(
                    system_prompt, user_prompt_template, text, model, temperature, max_tokens
                )
                
                # 记录处理信息
                total_tokens = self.count_tokens(system_prompt + user_prompt_template.format(text=text), model)
                self.logger.info(f"处理 {custom_id}: 输入 {total_tokens} tokens")
                
                # 调用API
                response = self.client.chat.completions.create(**api_params)
                
                # 处理响应
                return self._process_api_response(response, custom_id, model, source_file)
                
            except Exception as e:
                import traceback
                error_details = f"{type(e).__name__}: {str(e)}"
                self.logger.error(f"❌ {custom_id} 处理失败: {error_details}")
                self.logger.debug(f"完整错误堆栈: {traceback.format_exc()}")
                return self._create_error_result(custom_id, model, source_file, error_details)
    
    def _build_api_params(self, system_prompt: str, user_prompt_template: str, text: str, 
                         model: str, temperature: float, max_tokens: int) -> Dict[str, Any]:
        """构建API调用参数"""
        from config import get_token_param_name
        
        user_prompt = user_prompt_template.format(text=text)
        
        api_params = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": temperature
        }
        
        # 根据模型类型使用正确的token参数
        token_param = get_token_param_name(model)
        api_params[token_param] = max_tokens
        
        return api_params
    
    def _process_api_response(self, response, custom_id: str, model: str, source_file: str) -> Dict[str, Any]:
        """处理API响应"""
        try:
            # 处理不同类型的响应
            if isinstance(response, str):
                content = response
                usage_info = {"total_tokens": 0}
                model_name = model
            else:
                content = response.choices[0].message.content
                usage_info = {
                    "total_tokens": response.usage.total_tokens if response.usage else 0,
                    "prompt_tokens": response.usage.prompt_tokens if response.usage else 0,
                    "completion_tokens": response.usage.completion_tokens if response.usage else 0
                }
                model_name = response.model

            # 记录原始响应内容用于调试
            self.logger.debug(f"{custom_id} 原始API响应: {content[:500]}")

            # 解析JSON响应
            extracted_terms = self._parse_json_response(content)

            result = {
                "custom_id": custom_id,
                "extracted_terms": extracted_terms,
                "usage": usage_info,
                "model": model_name,
                "source_file": source_file,
                "created": int(time.time())
            }

            self.logger.info(f"✅ {custom_id} 处理完成: {usage_info.get('total_tokens', 0)} tokens")
            return result

        except Exception as e:
            import traceback
            self.logger.error(f"❌ {custom_id} 响应处理失败: {type(e).__name__}: {str(e)}")
            self.logger.debug(f"完整错误堆栈: {traceback.format_exc()}")
            # 返回错误结果而不是抛出异常
            return self._create_error_result(custom_id, model, source_file, f"响应处理失败: {type(e).__name__}: {str(e)}")
    
    def _parse_json_response(self, content: str) -> Dict[str, Any]:
        """解析JSON响应内容"""
        try:
            # 尝试直接解析（处理干净的JSON响应）
            parsed = json.loads(content)

            # 如果解析成功但返回的是字符串而不是对象，尝试再次解析
            if isinstance(parsed, str):
                self.logger.warning(f"API返回了JSON字符串而不是对象，尝试再次解析: {parsed[:100]}")
                parsed = json.loads(parsed)

            # 验证返回的是字典类型
            if not isinstance(parsed, dict):
                self.logger.warning(f"API返回了非字典类型: {type(parsed)}, 内容: {str(parsed)[:200]}")
                return {"raw_content": str(parsed)}

            # 验证是否包含terms字段
            if "terms" not in parsed:
                self.logger.warning(f"API响应缺少'terms'字段，返回的字段: {list(parsed.keys())}")
                # 如果响应中只有一个键，且其值是列表，可能就是terms
                if len(parsed) == 1:
                    single_key = list(parsed.keys())[0]
                    if isinstance(parsed[single_key], list):
                        self.logger.info(f"将字段'{single_key}'作为terms处理")
                        return {"terms": parsed[single_key]}
                # 否则将整个响应包装为raw_content
                return {"raw_content": str(parsed)}

            return parsed

        except json.JSONDecodeError as e:
            # JSON解析失败，尝试提取JSON块
            self.logger.warning(f"初始JSON解析失败: {str(e)}, 尝试提取JSON块")

            try:
                # 尝试查找JSON对象或数组
                json_start = content.find('{')
                if json_start == -1:
                    json_start = content.find('[')

                if json_start != -1:
                    # 查找对应的结束符
                    if content[json_start] == '{':
                        json_end = content.rfind('}')
                    else:
                        json_end = content.rfind(']')

                    if json_end != -1 and json_end > json_start:
                        extracted = content[json_start:json_end+1].strip()
                        self.logger.info(f"提取的JSON块: {extracted[:200]}...")
                        return json.loads(extracted)

                # 如果提取失败，记录详细错误
                self.logger.error(f"无法从响应中提取有效JSON，原始内容: {content[:500]}")
                return {"raw_content": content}

            except Exception as ex:
                self.logger.error(f"JSON块提取失败: {type(ex).__name__}: {str(ex)}")
                return {"raw_content": content}

        except Exception as e:
            # 捕获所有其他异常
            self.logger.error(f"JSON解析时发生意外错误: {type(e).__name__}: {str(e)}, 内容: {content[:500]}")
            return {"raw_content": content}
    
    def _create_error_result(self, custom_id: str, model: str, source_file: str, error_msg: str) -> Dict[str, Any]:
        """创建错误结果"""
        return {
            "custom_id": custom_id,
            "error": error_msg,
            "extracted_terms": {"raw_content": f"处理失败: {error_msg}"},
            "usage": {"total_tokens": 0},
            "model": model,
            "source_file": source_file,
            "created": int(time.time())
        }
    
    # =============================================================================
    # 批量并发处理
    # =============================================================================
    
    def process_batch_concurrent(self, 
                               texts: List[str],
                               system_prompt: str = None,
                               user_prompt_template: str = None,
                               model: str = "gpt-4-turbo-preview",
                               temperature: float = 0.1,
                               max_tokens: int = 4096,
                               max_concurrent: int = 10,
                               source_files: List[str] = None) -> List[Dict[str, Any]]:
        """
        并发处理批量文本
        
        Args:
            texts: 要处理的文本列表
            system_prompt: 系统提示词
            user_prompt_template: 用户提示词模板
            model: 模型名称
            temperature: 温度参数
            max_tokens: 最大输出token数
            max_concurrent: 最大并发数
            source_files: 来源文件名列表
            
        Returns:
            处理结果列表
        """
        self.logger.info(f"🚀 开始并发批处理: {len(texts)} 个文本，最大并发 {max_concurrent}")
        
        # 更新并发控制
        self.semaphore = threading.Semaphore(max_concurrent)
        
        results = []
        
        with ThreadPoolExecutor(max_workers=max_concurrent) as executor:
            # 提交所有任务
            future_to_id = self._submit_batch_tasks(
                executor, texts, system_prompt, user_prompt_template, 
                model, temperature, max_tokens, source_files
            )
            
            # 收集结果
            results = self._collect_batch_results(future_to_id, len(texts), model, source_files)
        
        # 按custom_id排序结果
        results.sort(key=lambda x: x.get("custom_id", ""))
        
        self.logger.info(f"✅ 并发批处理完成: {len(results)} 个结果")
        return results
    
    def _submit_batch_tasks(self, executor, texts: List[str], system_prompt: str, 
                           user_prompt_template: str, model: str, temperature: float, 
                           max_tokens: int, source_files: List[str]) -> Dict:
        """提交批处理任务"""
        future_to_id = {}
        
        for i, text in enumerate(texts):
            custom_id = f"term-extraction-{i+1}"
            source_file = source_files[i] if source_files and i < len(source_files) else None
            
            future = executor.submit(
                self.process_single_text,
                text=text,
                custom_id=custom_id,
                system_prompt=system_prompt,
                user_prompt_template=user_prompt_template,
                model=model,
                temperature=temperature,
                max_tokens=max_tokens,
                source_file=source_file
            )
            future_to_id[future] = custom_id
        
        return future_to_id
    
    def _collect_batch_results(self, future_to_id: Dict, total_count: int,
                              model: str, source_files: List[str]) -> List[Dict[str, Any]]:
        """收集批处理结果"""
        results = []
        completed_count = 0

        for future in as_completed(future_to_id):
            custom_id = future_to_id[future]
            try:
                result = future.result()
                results.append(result)
                completed_count += 1
                self.logger.info(f"📊 进度: {completed_count}/{total_count} 完成")

            except Exception as e:
                self.logger.error(f"❌ {custom_id} 处理异常: {e}")
                # 添加错误结果
                text_index = int(custom_id.split('-')[-1]) - 1
                source_file = source_files[text_index] if source_files and text_index < len(source_files) else f"text_{text_index+1}.txt"
                results.append(self._create_error_result(custom_id, model, source_file, str(e)))

        return results
    
    # =============================================================================
    # 术语去重和合并
    # =============================================================================
    
    def deduplicate_terms(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        去重和合并相同术语
        
        Args:
            results: 处理结果列表
            
        Returns:
            去重后的结果列表
        """
        from config import TERM_PROCESSING
        
        self.logger.info("开始术语去重处理...")
        
        # 收集所有术语
        all_terms = self._collect_all_terms(results)
        
        # 合并重复术语
        merged_terms, duplicate_count = self._merge_duplicate_terms(all_terms)
        
        # 创建合并结果
        merged_result = self._create_merged_result(results, merged_terms, duplicate_count)
        
        self.logger.info(f"去重完成: {len(merged_terms)} 个唯一术语 (移除 {duplicate_count} 个重复)")
        return [merged_result]
    
    def _collect_all_terms(self, results: List[Dict[str, Any]]) -> Dict[str, List[Dict]]:
        """收集所有术语（支持双语）"""
        from config import TERM_PROCESSING
        
        all_terms = {}  # {term_key: [term_objects]}
        case_sensitive = TERM_PROCESSING.get("case_sensitive_matching", False)
        
        for result in results:
            extracted_terms = result.get("extracted_terms", {})
            if "terms" in extracted_terms and isinstance(extracted_terms["terms"], list):
                for term in extracted_terms["terms"]:
                    if not isinstance(term, dict):
                        continue
                    
                    # 支持双语格式
                    eng_term = term.get("eng_term", "").strip()
                    zh_term = term.get("zh_term", "").strip()
                    
                    # 兼容旧格式
                    if not eng_term and not zh_term and "term" in term:
                        single_term = term["term"].strip()
                        term_key = single_term if case_sensitive else single_term.lower()
                        
                        if term_key not in all_terms:
                            all_terms[term_key] = []
                        all_terms[term_key].append({
                            "original_term": term["term"],
                            "source_id": result.get("custom_id", ""),
                            "source_file": result.get("source_file", "")
                        })
                    else:
                        # 双语格式：使用英文+中文组合作为key
                        eng_key = eng_term if case_sensitive else eng_term.lower()
                        zh_key = zh_term  # 中文不需要区分大小写
                        term_key = f"{eng_key}|{zh_key}"
                        
                        if term_key not in all_terms:
                            all_terms[term_key] = []
                        all_terms[term_key].append({
                            "original_eng_term": eng_term,
                            "original_zh_term": zh_term,
                            "source_id": result.get("custom_id", ""),
                            "source_file": result.get("source_file", "")
                        })
        
        return all_terms
    
    def _merge_duplicate_terms(self, all_terms: Dict[str, List[Dict]]) -> Tuple[List[Dict], int]:
        """合并重复术语"""
        merged_terms = []
        duplicate_count = 0
        
        for term_name, term_list in all_terms.items():
            if len(term_list) > 1:
                duplicate_count += len(term_list) - 1
            
            # 合并术语信息
            merged_term = self._merge_term_info(term_list)
            merged_terms.append(merged_term)
        
        return merged_terms, duplicate_count
    
    def _create_merged_result(self, results: List[Dict[str, Any]], merged_terms: List[Dict], 
                             duplicate_count: int) -> Dict[str, Any]:
        """创建合并结果"""
        return {
            "custom_id": "merged_terms",
            "extracted_terms": {
                "terms": merged_terms,
                "total_terms": len(merged_terms),
                "original_terms": sum(len(term_list) for term_list in self._collect_all_terms(results).values()),
                "duplicates_removed": duplicate_count
            },
            "usage": {"total_tokens": sum(r.get("usage", {}).get("total_tokens", 0) for r in results)},
            "model": "merged",
            "created": max(r.get("created", 0) for r in results) if results else 0
        }
    
    def _merge_term_info(self, term_list: List[Dict]) -> Dict:
        """
        合并相同术语的信息（支持双语）
        
        Args:
            term_list: 相同术语的列表
            
        Returns:
            合并后的术语信息
        """
        # 收集所有来源文件（去重）
        source_files = self._collect_source_files(term_list)
        
        # 判断是双语格式还是旧格式
        if "original_eng_term" in term_list[0]:
            # 双语格式
            if len(term_list) == 1:
                return {
                    "eng_term": term_list[0]["original_eng_term"],
                    "zh_term": term_list[0]["original_zh_term"],
                    "source_file": term_list[0].get("source_file", "")
                }
            else:
                # 选择最佳的术语版本
                best_eng_term = self._select_best_bilingual_term(term_list, "original_eng_term")
                best_zh_term = term_list[0]["original_zh_term"]  # 中文通常一致
                
                return {
                    "eng_term": best_eng_term,
                    "zh_term": best_zh_term,
                    "source_files": source_files if len(source_files) > 1 else source_files[0] if source_files else ""
                }
        else:
            # 旧格式（单一term字段）
            if len(term_list) == 1:
                return {
                    "term": term_list[0]["original_term"],
                    "source_file": term_list[0].get("source_file", "")
                }
            
            # 选择最佳的原始术语名（优先选择首字母大写的）
            best_term = self._select_best_term(term_list)
            
            return {
                "term": best_term,
                "source_files": source_files if len(source_files) > 1 else source_files[0] if source_files else ""
            }
    
    def _select_best_term(self, term_list: List[Dict]) -> str:
        """选择最佳的术语名称（旧格式）"""
        best_term = term_list[0]["original_term"]
        for term_info in term_list:
            if term_info["original_term"] and term_info["original_term"][0].isupper():
                best_term = term_info["original_term"]
                break
        return best_term
    
    def _select_best_bilingual_term(self, term_list: List[Dict], field_name: str) -> str:
        """选择最佳的双语术语名称（优先选择首字母大写的）"""
        best_term = term_list[0][field_name]
        for term_info in term_list:
            term_value = term_info.get(field_name, "")
            if term_value and term_value[0].isupper():
                best_term = term_value
                break
        return best_term
    
    def _collect_source_files(self, term_list: List[Dict]) -> List[str]:
        """收集来源文件列表"""
        return list(set([
            term_info.get("source_file", "") 
            for term_info in term_list 
            if term_info.get("source_file")
        ]))
    
    def _extract_source_filename(self, source_files: List[str]) -> str:
        """从源文件列表中提取主要文件名（不含扩展名和路径）"""
        if not source_files:
            return "unknown"
        
        # 取第一个文件作为主文件名
        main_file = source_files[0]
        
        # 处理虚拟文件名格式，如 "filename.pdf - 片段 1/5 (2621 tokens)"
        import re
        # 移除片段信息，只保留原始文件名部分
        clean_main_file = re.sub(r'\s*-\s*片段\s*\d+/\d+\s*\([^)]+\)', '', main_file)
        
        # 提取文件名部分，去掉路径和扩展名
        from pathlib import Path
        filename = Path(clean_main_file.strip()).stem
        
        # 清理文件名中的特殊字符，只保留中英文字母、数字和下划线
        clean_filename = re.sub(r'[^\w\u4e00-\u9fff]', '_', filename)
        
        return clean_filename[:50]  # 限制长度避免文件名过长
    
    def _count_total_terms(self, results: List[Dict[str, Any]]) -> int:
        """计算合并后的总术语数"""
        total = 0
        for result in results:
            if result.get("custom_id") == "merged_terms":
                extracted_terms = result.get("extracted_terms", {})
                if "terms" in extracted_terms:
                    total = len(extracted_terms["terms"])
                elif "total_terms" in extracted_terms:
                    total = extracted_terms["total_terms"]
                break
        
        return total
    
    # =============================================================================
    # 结果保存
    # =============================================================================
    
    def save_processed_results(self, 
                              results: List[Dict[str, Any]], 
                              output_format: str = "json",
                              source_filename: str = "",
                              model_name: str = "",
                              total_terms: int = 0) -> str:
        """
        保存处理结果到文件
        
        Args:
            results: 处理结果列表
            output_format: 输出格式 (json/csv/txt/excel)
            source_filename: 源文件名（不含扩展名）
            model_name: 使用的模型名称
            total_terms: 合并后的总术语数
            
        Returns:
            输出文件路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 构建新的命名格式: 文件-日期-模型-总条数
        if source_filename and model_name and total_terms > 0:
            filename_base = f"{source_filename}_{timestamp}_{model_name}_{total_terms}terms"
        else:
            filename_base = f"processed_terms_{timestamp}"
        
        if output_format == "json":
            return self._save_json_results(results, filename_base)
        elif output_format == "csv":
            return self._save_csv_results(results, filename_base)
        elif output_format == "txt":
            return self._save_txt_results(results, filename_base)
        elif output_format == "excel":
            return self._save_excel_results(results, filename_base)
        elif output_format == "tbx":
            return self._save_tbx_results(results, filename_base)
        else:
            raise ValueError(f"不支持的输出格式: {output_format}")
    
    def _save_json_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """保存JSON格式结果"""
        output_file = self.base_dir / f"{filename_base}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"结果已保存到: {output_file}")
        return str(output_file)
    
    def _save_csv_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """保存CSV格式结果（支持双语）"""
        import csv
        
        output_file = self.base_dir / f"{filename_base}.csv"
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:  # 使用utf-8-sig支持Excel
            if results:
                writer = csv.writer(f)
                # 写入标题行（双语）
                writer.writerow(["custom_id", "英文术语", "中文术语", "来源文件", "模型", "token使用"])
                
                for result in results:
                    terms = result.get("extracted_terms", {})
                    if "terms" in terms and isinstance(terms["terms"], list):
                        for term in terms["terms"]:
                            # 支持双语格式
                            eng_term = term.get("eng_term", "")
                            zh_term = term.get("zh_term", "")
                            
                            # 兼容旧格式（单一term字段）
                            if not eng_term and not zh_term:
                                single_term = term.get("term", "")
                                # 简单判断：如果包含中文字符，放入中文列
                                if any('\u4e00' <= char <= '\u9fff' for char in single_term):
                                    zh_term = single_term
                                else:
                                    eng_term = single_term
                            
                            writer.writerow([
                                result.get("custom_id", ""),
                                eng_term,
                                zh_term,
                                term.get("source_file", result.get("source_file", "")),
                                result.get("model", ""),
                                result.get("usage", {}).get("total_tokens", 0)
                            ])
                    else:
                        # 兼容其他格式
                        writer.writerow([
                            result.get("custom_id", ""),
                            "",
                            str(terms.get("raw_content", "")),
                            result.get("source_file", ""),
                            result.get("model", ""),
                            result.get("usage", {}).get("total_tokens", 0)
                        ])
        
        self.logger.info(f"结果已保存到: {output_file}")
        return str(output_file)
    
    def _save_txt_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """保存TXT格式结果（支持双语）"""
        output_file = self.base_dir / f"{filename_base}.txt"
        with open(output_file, 'w', encoding='utf-8') as f:
            for i, result in enumerate(results, 1):
                f.write(f"=== 结果 {i}: {result.get('custom_id', 'unknown')} ===\n")
                terms = result.get("extracted_terms", {})
                
                if "terms" in terms and isinstance(terms["terms"], list):
                    f.write(f"提取的术语数量: {len(terms['terms'])}\n")
                    f.write(f"来源文件: {result.get('source_file', '未知')}\n")
                    f.write(f"使用模型: {result.get('model', '未知')}\n")
                    f.write(f"Token使用: {result.get('usage', {}).get('total_tokens', 0)}\n\n")
                    
                    for j, term in enumerate(terms["terms"], 1):
                        # 支持双语格式
                        eng_term = term.get("eng_term", "")
                        zh_term = term.get("zh_term", "")
                        
                        # 兼容旧格式
                        if not eng_term and not zh_term:
                            single_term = term.get("term", "")
                            f.write(f"{j}. {single_term}\n")
                        else:
                            # 双语格式输出
                            f.write(f"{j}. {eng_term}\n")
                            f.write(f"   {zh_term}\n")
                        
                        if term.get('source_file'):
                            f.write(f"   来源: {term.get('source_file')}\n")
                else:
                    f.write(f"原始内容: {terms.get('raw_content', '无内容')}\n")
                
                f.write("="*50 + "\n\n")
        
        self.logger.info(f"结果已保存到: {output_file}")
        return str(output_file)
    
    def _save_excel_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """保存Excel格式结果"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        output_file = self.base_dir / f"{filename_base}.xlsx"
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "术语抽取结果"
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 写入标题行（双语）
        headers = ["序号", "英文术语", "中文术语", "来源文件", "模型", "Token使用", "处理时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        row_num = 2
        term_count = 0
        
        for result in results:
            terms = result.get("extracted_terms", {})
            if "terms" in terms and isinstance(terms["terms"], list):
                for term in terms["terms"]:
                    term_count += 1
                    
                    # 处理来源文件信息
                    source_file = ""
                    if term.get('source_files'):
                        if isinstance(term['source_files'], list):
                            source_file = "; ".join(term['source_files'])
                        else:
                            source_file = str(term['source_files'])
                    elif term.get('source_file'):
                        source_file = term['source_file']
                    elif result.get('source_file'):
                        source_file = result['source_file']
                    
                    # 处理时间
                    created_time = ""
                    if result.get('created'):
                        created_time = datetime.fromtimestamp(result['created']).strftime("%Y-%m-%d %H:%M:%S")
                    
                    # 支持双语格式
                    eng_term = term.get("eng_term", "")
                    zh_term = term.get("zh_term", "")
                    
                    # 兼容旧格式（单一term字段）
                    if not eng_term and not zh_term:
                        single_term = term.get("term", "")
                        # 简单判断：如果包含中文字符，放入中文列
                        if any('\u4e00' <= char <= '\u9fff' for char in single_term):
                            zh_term = single_term
                        else:
                            eng_term = single_term
                    
                    # 写入行数据
                    row_data = [
                        term_count,  # 序号
                        eng_term,  # 英文术语
                        zh_term,  # 中文术语
                        source_file,  # 来源文件
                        result.get('model', ''),  # 模型
                        result.get('usage', {}).get('total_tokens', 0),  # Token使用
                        created_time  # 处理时间
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.border = border
                        # 居中对齐序号和Token使用列
                        if col in [1, 6]:
                            cell.alignment = Alignment(horizontal="center")
                    
                    row_num += 1
            else:
                # 处理其他格式的数据
                term_count += 1
                raw_content = terms.get('raw_content', '无内容')
                
                row_data = [
                    term_count,
                    raw_content[:100] + "..." if len(raw_content) > 100 else raw_content,
                    result.get('source_file', ''),
                    result.get('model', ''),
                    result.get('usage', {}).get('total_tokens', 0),
                    ""
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border
                    if col in [1, 5]:
                        cell.alignment = Alignment(horizontal="center")
                
                row_num += 1
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            # 设置列宽，最小10，最大50
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 添加统计信息工作表
        stats_ws = wb.create_sheet("统计信息")
        stats_data = [
            ["统计项目", "数值"],
            ["总术语数", term_count],
            ["处理结果数", len(results)],
            ["总Token使用", sum(r.get('usage', {}).get('total_tokens', 0) for r in results)],
            ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = stats_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 标题行
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                cell.border = border
        
        # 调整统计信息表列宽
        for col in range(1, 3):
            column_letter = get_column_letter(col)
            stats_ws.column_dimensions[column_letter].width = 15
        
        # 保存文件
        wb.save(output_file)
        
        self.logger.info(f"Excel结果已保存到: {output_file}")
        return str(output_file)
    
    def _save_tbx_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """保存TBX格式结果"""
        import xml.etree.ElementTree as ET
        from xml.dom import minidom
        import re
        
        output_file = self.base_dir / f"{filename_base}.tbx"
        
        # 智能检测主要语言
        primary_language = self._detect_primary_language(results)
        
        # 创建TBX根元素
        root = ET.Element("tbx", attrib={
            "type": "TBX-Default",
            "style": "dct",
            "xml:lang": primary_language,
            "xmlns": "urn:iso:std:iso:30042:ed-2"
        })
        
        # 添加TBX头部信息
        tbx_header = ET.SubElement(root, "tbxHeader")
        
        # 文件描述
        file_desc = ET.SubElement(tbx_header, "fileDesc")
        title_stmt = ET.SubElement(file_desc, "titleStmt")
        title = ET.SubElement(title_stmt, "title")
        title.text = "军事航天术语库"
        
        # 发布信息
        pub_stmt = ET.SubElement(file_desc, "publicationStmt")
        publisher = ET.SubElement(pub_stmt, "p")
        publisher.text = f"Generated by LLM Term Extraction Tool on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 源描述
        source_desc = ET.SubElement(file_desc, "sourceDesc")
        source_p = ET.SubElement(source_desc, "p")
        source_p.text = "Extracted from military and aerospace documents using AI processing"
        
        # 编码描述
        encoding_desc = ET.SubElement(tbx_header, "encodingDesc")
        encoding_p = ET.SubElement(encoding_desc, "p", attrib={"type": "XCSURI"})
        encoding_p.text = "http://www.lisa.org/TBX-Specification.33.0.html"
        
        # 修订描述
        revision_desc = ET.SubElement(tbx_header, "revisionDesc")
        change = ET.SubElement(revision_desc, "change")
        change_date = ET.SubElement(change, "date")
        change_date.text = datetime.now().strftime("%Y-%m-%d")
        change_resp = ET.SubElement(change, "respName")
        change_resp.text = "LLM Term Extraction Tool"
        change_item = ET.SubElement(change, "item", attrib={"type": "create"})
        change_item.text = "Initial term extraction"
        
        # 创建术语体
        text_body = ET.SubElement(root, "text")
        body = ET.SubElement(text_body, "body")
        
        # 处理术语数据
        term_count = 0
        for result in results:
            terms = result.get("extracted_terms", {})
            if "terms" in terms and isinstance(terms["terms"], list):
                for term in terms["terms"]:
                    term_count += 1
                    
                    # 创建术语条目
                    term_entry = ET.SubElement(body, "termEntry", attrib={"id": f"term_{term_count}"})
                    
                    # 添加管理信息
                    admin_grp = ET.SubElement(term_entry, "adminGrp")
                    admin = ET.SubElement(admin_grp, "admin", attrib={"type": "subjectField"})
                    admin.text = "military_aerospace"
                    
                    # 处理来源文件信息
                    source_file = ""
                    if term.get('source_files'):
                        if isinstance(term['source_files'], list):
                            source_file = "; ".join(term['source_files'])
                        else:
                            source_file = str(term['source_files'])
                    elif term.get('source_file'):
                        source_file = term['source_file']
                    elif result.get('source_file'):
                        source_file = result['source_file']
                    
                    if source_file:
                        source_admin = ET.SubElement(admin_grp, "admin", attrib={"type": "source"})
                        source_admin.text = source_file
                    
                    # 添加处理信息
                    process_admin = ET.SubElement(admin_grp, "admin", attrib={"type": "processStatus"})
                    process_admin.text = "provisionallyProcessed"
                    
                    # 支持双语格式
                    eng_term = term.get("eng_term", "")
                    zh_term = term.get("zh_term", "")
                    
                    # 兼容旧格式
                    if not eng_term and not zh_term:
                        single_term = term.get('term', '')
                        term_language = self._detect_term_language(single_term)
                        
                        # 创建单一语言组
                        lang_grp = ET.SubElement(term_entry, "langGrp", attrib={"xml:lang": term_language})
                        term_grp = ET.SubElement(lang_grp, "termGrp")
                        term_elem = ET.SubElement(term_grp, "term")
                        term_elem.text = single_term
                    else:
                        # 双语格式：创建英文语言组
                        if eng_term:
                            lang_grp_en = ET.SubElement(term_entry, "langGrp", attrib={"xml:lang": "en"})
                            term_grp_en = ET.SubElement(lang_grp_en, "termGrp")
                            term_elem_en = ET.SubElement(term_grp_en, "term")
                            term_elem_en.text = eng_term
                        
                        # 双语格式：创建中文语言组
                        if zh_term:
                            lang_grp_zh = ET.SubElement(term_entry, "langGrp", attrib={"xml:lang": "zh"})
                            term_grp_zh = ET.SubElement(lang_grp_zh, "termGrp")
                            term_elem_zh = ET.SubElement(term_grp_zh, "term")
                            term_elem_zh.text = zh_term
                    
                    # 添加提取时间（在最后一个termGrp中）
                    if result.get('created'):
                        # 获取最后创建的termGrp
                        last_lang_grp = list(term_entry.findall("langGrp"))[-1] if term_entry.findall("langGrp") else None
                        if last_lang_grp is not None:
                            last_term_grp = last_lang_grp.find("termGrp")
                            if last_term_grp is not None:
                                date_admin = ET.SubElement(last_term_grp, "admin", attrib={"type": "created"})
                                date_admin.text = datetime.fromtimestamp(result['created']).strftime("%Y-%m-%d")
        
        # 格式化XML并保存
        rough_string = ET.tostring(root, encoding='unicode')
        reparsed = minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="  ", encoding=None)
        
        # 移除空行
        pretty_lines = [line for line in pretty_xml.split('\n') if line.strip()]
        final_xml = '\n'.join(pretty_lines)
        
        # 添加XML声明和DOCTYPE
        xml_declaration = '<?xml version="1.0" encoding="UTF-8"?>\n'
        doctype = '<!DOCTYPE tbx SYSTEM "TBXcoreStructV02.dtd">\n'
        final_content = xml_declaration + doctype + '\n'.join(final_xml.split('\n')[1:])
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(final_content)
        
        self.logger.info(f"TBX结果已保存到: {output_file}")
        return str(output_file)
    
    def _detect_primary_language(self, results: List[Dict[str, Any]]) -> str:
        """检测术语的主要语言"""
        import re
        
        chinese_count = 0
        english_count = 0
        total_terms = 0
        
        for result in results:
            terms = result.get("extracted_terms", {})
            if "terms" in terms and isinstance(terms["terms"], list):
                for term in terms["terms"]:
                    term_text = term.get('term', '')
                    if term_text:
                        total_terms += 1
                        # 优先检测中文，如果包含中文字符就算中文术语
                        if re.search(r'[\u4e00-\u9fff]', term_text):  # 包含中文字符
                            chinese_count += 1
                        # 只有纯英文才算英文术语
                        elif re.search(r'^[a-zA-Z0-9\s\-\(\)\.]+$', term_text):  # 纯英文字符
                            english_count += 1
        
        # 根据主要语言返回语言代码
        if chinese_count > english_count:
            return "zh-CN"
        elif english_count > 0:
            return "en"
        else:
            return "zh-CN"  # 默认中文
    
    def _detect_term_language(self, term_text: str) -> str:
        """检测单个术语的语言"""
        import re
        
        if not term_text:
            return "zh-CN"
        
        # 检查是否包含中文字符
        if re.search(r'[\u4e00-\u9fff]', term_text):
            return "zh-CN"
        # 检查是否主要是英文字符
        elif re.search(r'[a-zA-Z]', term_text):
            return "en"
        else:
            return "zh-CN"  # 默认中文
    
    # =============================================================================
    # 完整处理流程
    # =============================================================================
    
    def run_extraction_only(self, 
                            texts: List[str],
                            system_prompt: str = None,
                            user_prompt_template: str = None,
                            model: str = "gpt-4-turbo-preview",
                            temperature: float = 0.1,
                            max_tokens: int = 4096,
                            max_concurrent: int = 10,
                            description: str = "术语抽取批处理任务",
                            source_files: List[str] = None) -> Dict[str, Any]:
        """
        只运行抽取流程，不保存最终结果文件
        
        Args:
            texts: 要处理的文本列表
            system_prompt: 系统提示词
            user_prompt_template: 用户提示词模板
            model: 模型名称
            temperature: 温度参数
            max_tokens: 最大输出token数
            max_concurrent: 最大并发数
            description: 任务描述
            source_files: 来源文件名列表
            
        Returns:
            Dict[str, Any]: 包含处理结果的字典
        """
        self.logger.info("🚀 开始术语抽取流程")

        try:
            # 步骤1: 并发处理所有文本
            results = self._run_concurrent_processing(
                texts, system_prompt, user_prompt_template, model,
                temperature, max_tokens, max_concurrent, source_files
            )

            # 步骤2: 保存原始结果
            raw_file = self._save_raw_results(results)

            # 步骤3: 术语去重处理
            merged_results = self._run_deduplication(results)

            return {
                "raw_results": results,
                "merged_results": merged_results,
                "raw_file": raw_file
            }

        except Exception as e:
            self.logger.error(f"❌ 抽取流程执行失败: {e}")
            raise

    def run_complete_pipeline(self, 
                             texts: List[str],
                             system_prompt: str = None,
                             user_prompt_template: str = None,
                             model: str = "gpt-4-turbo-preview",
                             temperature: float = 0.1,
                             max_tokens: int = 4096,
                             max_concurrent: int = 10,
                             description: str = "术语抽取批处理任务",
                             output_format: str = "json",
                             source_files: List[str] = None) -> Dict[str, str]:
        """
        运行完整的批处理流程
        
        Args:
            texts: 要处理的文本列表
            system_prompt: 系统提示词
            user_prompt_template: 用户提示词模板
            model: 模型名称
            temperature: 温度参数
            max_tokens: 最大输出token数
            max_concurrent: 最大并发数
            description: 任务描述
            output_format: 输出格式
            source_files: 来源文件名列表
            
        Returns:
            Dict[str, str]: 包含各个文件路径的字典
        """
        self.logger.info("🚀 开始完整LLM批处理流程")

        try:
            # 步骤1: 并发处理所有文本
            results = self._run_concurrent_processing(
                texts, system_prompt, user_prompt_template, model,
                temperature, max_tokens, max_concurrent, source_files
            )
            
            # 步骤2: 保存原始结果
            raw_file = self._save_raw_results(results)
            
            # 步骤3: 术语去重处理
            merged_results = self._run_deduplication(results)
            
            # 步骤4: 保存最终结果
            # 提取元数据用于文件命名
            source_filename = self._extract_source_filename(source_files)
            model_name = model.replace("-", "").replace(".", "")  # 清理模型名用于文件名
            total_terms = self._count_total_terms(merged_results)
            
            processed_file = self.save_processed_results(
                merged_results, 
                output_format,
                source_filename,
                model_name,
                total_terms
            )
            
            return {
                "raw_results": raw_file,
                "processed_results": processed_file
            }
            
        except Exception as e:
            self.logger.error(f"❌ 完整流程执行失败: {e}")
            raise
    
    def _run_concurrent_processing(self, texts: List[str], system_prompt: str, 
                                  user_prompt_template: str, model: str, temperature: float, 
                                  max_tokens: int, max_concurrent: int, 
                                  source_files: List[str]) -> List[Dict[str, Any]]:
        """运行并发处理"""
        self.logger.info("🔄 步骤1: 并发处理文本")
        return self.process_batch_concurrent(
            texts=texts,
            system_prompt=system_prompt,
            user_prompt_template=user_prompt_template,
            model=model,
            temperature=temperature,
            max_tokens=max_tokens,
            max_concurrent=max_concurrent,
            source_files=source_files
        )
    
    def _save_raw_results(self, results: List[Dict[str, Any]]) -> str:
        """保存原始结果"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        raw_result_file = self.base_dir / f"raw_results_{timestamp}.json"
        with open(raw_result_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"原始结果已保存到: {raw_result_file}")
        return str(raw_result_file)
    
    def _run_deduplication(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """运行去重处理"""
        self.logger.info("🔄 步骤2: 术语去重处理")
        return self.deduplicate_terms(results)


# =============================================================================
# 文本加载和处理工具函数
# =============================================================================

def _save_intermediate_text(file_path: str, text: str):
    """保存中间提取的文本文件"""
    import os
    from pathlib import Path
    
    # 创建extracted_texts文件夹
    extracted_dir = Path("extracted_texts")
    extracted_dir.mkdir(exist_ok=True)
    
    # 生成输出文件名
    source_file = Path(file_path)
    output_file = extracted_dir / f"{source_file.stem}.txt"
    
    try:
        # 保存提取的文本
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(text)
        
        print(f"💾 中间文本已保存: {output_file}")
        
    except Exception as e:
        print(f"⚠️  保存中间文本失败: {e}")

def load_texts_from_file(file_path: str,
                        chunk_size: Optional[int] = None,
                        use_smart_splitter: bool = True,
                        overlap_size: int = 200) -> List[str]:
    """
    从文件加载文本并进行智能分割

    Args:
        file_path: 文件路径
        chunk_size: 分块大小（字符数），None表示不分块
        use_smart_splitter: 是否使用智能分割器
        overlap_size: 重叠大小（字符数）

    Returns:
        分割后的文本列表
    """
    from file_processor import FileProcessor
    from text_splitter import TextSplitter
    import os

    # 加载文件内容
    processor = FileProcessor()
    
    try:
        if file_path.endswith('.pdf'):
            texts = processor.extract_pdf_text(file_path)
        elif file_path.endswith(('.docx', '.doc')):
            texts = processor.extract_docx_text(file_path)
        else:
            # 文本文件
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            texts = [content]
        
        if not texts or not any(text.strip() for text in texts):
            raise ValueError("文件内容为空或无法提取")
        
        # 合并所有文本
        full_text = '\n\n'.join(texts)
        
        # 保存中间文本文件到extracted_texts文件夹
        _save_intermediate_text(file_path, full_text)
        
        # 根据配置进行分割
        if chunk_size and use_smart_splitter:
            # 使用智能分割器
            max_tokens = max(chunk_size // 4, 500)  # 最小500 tokens
            overlap_tokens = min(overlap_size // 4, max_tokens // 10)  # 重叠不超过10%
            
            splitter = TextSplitter(
                max_tokens=max_tokens,
                overlap_tokens=overlap_tokens
            )
            return splitter.split_text_with_metadata(full_text, Path(file_path).name)
        elif chunk_size:
            # 简单按段落分割
            splitter = TextSplitter(max_tokens=10000)  # 设置很大的值，避免合并
            chunks = splitter.split_by_paragraphs(full_text)
            
            # 添加文件标识
            labeled_chunks = []
            for i, chunk in enumerate(chunks, 1):
                if len(chunks) > 1:
                    labeled_chunk = f"[文件: {Path(file_path).name} - 片段 {i}/{len(chunks)} ({len(chunk)} 字符)]\n{chunk}"
                else:
                    labeled_chunk = f"[文件: {Path(file_path).name}]\n{chunk}"
                labeled_chunks.append(labeled_chunk)
            
            return labeled_chunks
        else:
            # 不分割，返回完整文本
            return [f"[文件: {Path(file_path).name}]\n{full_text}"]
        
    except Exception as e:
        raise ValueError(f"加载文件失败 {file_path}: {e}")


if __name__ == "__main__":
    # 简单测试
    print("LLM处理器模块加载成功")